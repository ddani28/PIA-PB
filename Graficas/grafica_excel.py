import requests
import openpyxl
from openpyxl.chart import BarChart, Reference



ciudad = "Monterrey"
pais = "MX"


def graficar_excel(ciudad,pais):
    url = "http://api.openweathermap.org/data/2.5/forecast"
    apikey = "33f8a5645d6c546cbf65483583c7149a"  
    params = {
        "q": f"{ciudad},{pais}",
        "appid": apikey,
        "units": "metric"  
    }

    response = requests.get(url, params=params)
    data = response.json()
    horas = [dato['dt_txt'] for dato in data['list']]
    temperaturas = [dato['main']['temp'] for dato in data['list']]
    wb = openpyxl.Workbook()
    hoja = wb.active
    hoja.cell(row=1, column=1, value="Hora")
    hoja.cell(row=1, column=2, value="Temperatura (°C)")

    for fila, (hora, temperatura) in enumerate(zip(horas, temperaturas), start=2):
        hoja.cell(row=fila, column=1, value=hora)
        hoja.cell(row=fila, column=2, value=temperatura)
    grafico = BarChart()
    grafico.title = "Pronóstico del clima en Monterrey"
    grafico.x_axis.title = "Hora"
    grafico.y_axis.title = "Temperatura (°C)"
    categorias = Reference(hoja, min_col=1, min_row=2, max_row=len(horas) + 1)
    valores = Reference(hoja, min_col=2, min_row=1, max_row=len(temperaturas) + 1)
    grafico.add_data(valores, titles_from_data=True)
    grafico.set_categories(categorias)
    hoja.add_chart(grafico, "D2")
    nombre_archivo_excel = "reporte_clima.xlsx"
    wb.save(nombre_archivo_excel)
    print(f"El gráfico ha sido creado y guardado en '{nombre_archivo_excel}'.")
